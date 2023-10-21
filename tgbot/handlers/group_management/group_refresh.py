from aiogram import types
from aiogram.dispatcher import FSMContext
import openpyxl
import io
from urllib.request import urlopen
from transliterate import translit
import itertools
import random

import tgbot.keyboards.reply as rkb
from tgbot.models.db import Database
from tgbot.filters.user_type import UserTypeFilter
from tgbot.misc.states import GroupRefreshStates
from tgbot.models.database_instance import db

group_name_request = "<b>Введите название группы, которую хотите обновить</b>"


async def get_group_name(message: types.Message, state: FSMContext):
    groups = await db.get_group_names()
    if len(groups) == 0:
        await message.answer(text="Группы еще не добавлены", reply_markup=rkb.manager_keyboard)
    else:
        await state.update_data(group_list=groups)
        group_msg = ""
        for group in groups:
            group_msg += f"◦ `{group[0]}`;\n"
        # group_msg = sorted(group_msg)
        await message.answer(text=f"*Доступные группы:*\n{group_msg}\n"
                                  f"*Введите название группы, которую хотите обновить*",
                             reply_markup=rkb.group_name_cancel_keyboard, parse_mode="MARKDOWN")
        await GroupRefreshStates.refresh_group_state.set()


async def check_group_name(message: types.Message, state: FSMContext):
    bot = message.bot
    chat_id = message.from_user.id
    group_name = message.text
    group_name_list = await db.get_group_names()
    group_exist = False
    for group in group_name_list:
        if group[0] == group_name:
            group_exist = True
            await state.update_data(group_name=group_name)
            await refresh_students(bot, chat_id, state)
    if not group_exist:
        await bot.send_message(text="<b>Такой группы не существует</b>", chat_id=chat_id,
                               reply_markup=rkb.manager_keyboard)
        await state.finish()


async def refresh_students(bot, chat_id, state: FSMContext):
    data = await state.get_data()
    group_name = data.get("group_name")
    link = await db.get_group_url(group_name)
    for l in link:
        link = l
    file_id = link.split('/')[-2]
    file_url = f'https://drive.google.com/u/0/uc?id={file_id}&export=download'
    wb = openpyxl.load_workbook(filename=io.BytesIO(urlopen(file_url).read()))
    sheet = wb.active

    target_row = 0
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value is not None:
                target_row = cell.row
                break
        else:
            continue
        break

    for cell in sheet[target_row]:
        if cell.value == '№':
            column_letter = cell.column_letter
            column_values = sheet[column_letter]
            for column_cell in column_values:
                if column_cell.value == 1:
                    target_row = column_cell.row
            break

    learning = set()
    not_learning = set()
    for row in sheet.iter_rows(min_row=target_row, values_only=True):
        if row[5] is not None:
            last_name = row[5].strip()
        else:
            last_name = row[5]
        if row[6] is not None:
            first_name = row[6].strip()
        else:
            first_name = row[6]
        if row[7] is not None:
            middle_name = row[7].strip()
        else:
            middle_name = row[7]
        if all((last_name, first_name)):
            fio = (last_name, first_name, middle_name)
            if row[0] is not None and row[1] == "учится":
                learning.add(fio)
            elif row[0] is None and row[1] != "учится":
                not_learning.add(fio)

    students_data = ""
    year = group_name[-2:]
    students = await db.get_students_by_group_name(group_name)  # !!!
    student_list = []
    for student in students:
        if student[2] is not None:
            student_list.append(f"{student[0]} {student[1]} {student[2]}")
        else:
            student_list.append(f"{student[0]} {student[1]}")
    for fio in learning:
        if " ".join(fio) not in student_list:
            user = []
            for name in fio:
                user.append(name)
            if len(user) < 3:
                user.append(None)
            login = await generate_login(user, year)

            password = random.randint(100000, 999999)
            await db.add_user(list(fio), login, password, "student", group_name)
            if fio[2] is not None:
                students_data += fio[0] + " " + fio[1] + " " + fio[2] + " " + login + " " + str(password) + "\n"
            else:
                students_data += fio[0] + " " + fio[1] + " " + login + " " + str(password) + "\n"
    deleted_students = ""
    for fio in not_learning:
        if " ".join(fio) in student_list:
            await db.delete_user(fio)
            if fio[2] is not None:
                deleted_students += f"{fio[0]} {fio[1]} {fio[2]}"
            else:
                deleted_students += f"{fio[0]} {fio[1]}"
    await bot.send_message(chat_id=chat_id,
                           text=f"Группа {group_name} обновлена\n"
                                f"<b>Данные новых студентов:</b> \n{students_data}\n\n"
                                f"<b>Удаленные студенты:</b>\n{deleted_students}",
                           reply_markup=rkb.manager_keyboard)
    await state.finish()


async def generate_login(fio, year):
    first_name_en = translit(fio[0], 'ru', reversed=True).lower()
    last_name_en = translit(fio[1], 'ru', reversed=True).lower()
    if fio[2] is None:
        initials = first_name_en[0] + last_name_en[0]
    else:
        middle_name_en = translit(fio[2], 'ru', reversed=True).lower()
        initials = first_name_en[0] + last_name_en[0] + middle_name_en[0]
    login_base = initials + "." + year

    logins = await db.get_user_logins()
    for i in itertools.count(start=1):
        login = f"{login_base}.{str(i)}"
        if login not in [log[0] for log in logins]:
            return login

        logins.append((login,))


def register_group_refresh(dp):
    dp.register_message_handler(get_group_name, UserTypeFilter("manager"), content_types=['text'],
                                text='Обновить группу')
    dp.register_message_handler(check_group_name, state=GroupRefreshStates.refresh_group_state)
